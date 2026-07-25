# -*- coding: utf-8 -*-
"""對抗性稽核：合併時，官方題庫與社群題庫對同一題是否給出不同答案？

如果有，目前的 merge() 會**靜默保留官方版、丟棄社群版** ——
但兩邊答案不一致本身就是一個訊號：其中一方是錯的，或該題已因法規修正而變動。
把它默默丟掉，等於銷毀了一條「這題有爭議」的證據。
"""
import csv, os, re, sys, unicodedata
from collections import Counter
sys.stdout.reconfigure(encoding="utf-8")

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
COLLECTION = os.path.dirname(REPO)
OFFICIAL_CSV = os.path.join(COLLECTION, "01_官方題庫_票券公會", "票券商業務人員_官方參考題庫_480題.csv")
COMMUNITY_XLSX = os.path.join(COLLECTION, "05_參考資料", "票券商業務人員_考古題整理_940題.xlsx")
KEYS = ["A", "B", "C", "D"]


def norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[，。？?、：:；;（）()「」『』【】\[\].,‐-―\-_/\\~｜|]", "", s)
    return s.replace("臺", "台")


def clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def key(a):
    a = clean(a)
    return KEYS[int(a) - 1] if a in ("1", "2", "3", "4") else None


official = {}
with open(OFFICIAL_CSV, encoding="utf-8") as fh:
    for r in list(csv.reader(fh))[1:]:
        if len(r) < 7 or not clean(r[2]):
            continue
        k = key(r[1])
        if k:
            official[norm(r[2])] = (int(float(r[0])), k, clean(r[2]),
                                    [clean(x) for x in r[3:7]])

from openpyxl import load_workbook
wb = load_workbook(COMMUNITY_XLSX, data_only=True)
conflicts, agree, seen = [], 0, set()
for ws in wb.worksheets:
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[5]:
            continue
        k = key(row[3])
        if not k:
            continue
        fp = norm(row[5])
        if fp not in official or fp in seen:
            continue
        seen.add(fp)
        ono, oans, ostem, oopts = official[fp]
        if oans == k:
            agree += 1
        else:
            conflicts.append((ono, oans, k, ostem[:60], oopts))

print(f"官方 ∩ 社群 的重複題：{agree + len(conflicts)} 題")
print(f"  兩邊答案一致：{agree}")
print(f"  兩邊答案衝突：{len(conflicts)}   ← 目前被靜默丟棄")
if conflicts:
    print("\n=== 衝突明細 ===")
    for ono, oans, cans, stem, opts in conflicts:
        print(f"\n[官方 #{ono}] 官方答案={oans}　社群答案={cans}")
        print(f"  {stem}…")
        for i, o in enumerate(opts):
            mark = ""
            if KEYS[i] == oans:
                mark += " ←官方"
            if KEYS[i] == cans:
                mark += " ←社群"
            print(f"    ({KEYS[i]}) {o[:52]}{mark}")
