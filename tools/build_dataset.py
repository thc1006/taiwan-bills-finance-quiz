#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
從原始蒐集檔建置 quiz-app 的題庫資料集。

輸入（皆為本次調研蒐集的原始檔，位於本 repo 之外的資料夾）：
  1. 票券公會官方參考題庫 480 題  (CSV)
  2. 網路流傳考古題整理 940 題    (XLSX，兩個工作表：考題整理 / 計算題)

輸出：
  quiz-app/src/data/dataset.json

設計原則 —— 誠實的 provenance：
  這支腳本**不會**宣稱任何題目的答案「已查證」。它只記錄每題**來自哪裡**。
  官方公會題庫的答案來自公會釋出的檔案；社群整理題庫的答案來自不具名的網路彙整，
  兩者都可能因法規修正而過時。答案正確性的實際查證由 check_law_citations.py
  以「引用法條是否仍存在」為代理指標另行標記，且該指標**不等於**答案正確。
"""
from __future__ import annotations

import csv
import json
import os
import re
import sys
import unicodedata
from datetime import datetime, timezone

sys.stdout.reconfigure(encoding="utf-8")

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
COLLECTION = os.path.dirname(REPO)  # 蒐集資料的根目錄（票券商業務人員/）

OFFICIAL_CSV = os.path.join(
    COLLECTION, "01_官方題庫_票券公會", "票券商業務人員_官方參考題庫_480題.csv"
)
COMMUNITY_XLSX = os.path.join(
    COLLECTION, "05_參考資料", "票券商業務人員_考古題整理_940題.xlsx"
)
OUT = os.path.join(REPO, "quiz-app", "src", "data", "dataset.json")

KEYS = ["A", "B", "C", "D"]
SUBJ_LAW = "票券金融法規"
SUBJ_PRACTICE = "票券金融實務"


# ---------------------------------------------------------------- helpers
def norm_text(s) -> str:
    """題目指紋：NFKC 正規化後剝除空白與標點。

    與 quiz-app/src/utils/question-identity.ts 的 fingerprint() 必須維持同一套定義。
    兩邊定義若不一致，CI 擋得住的重複題，使用者的考卷上照樣會出現兩次。
    """
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = re.sub(r"\s+", "", s)
    # ‐-― 一次涵蓋 ‐ ‑ ‒ – — ―。來源檔混用了其中多種，
    # 只列 - — – 會漏掉 U+2010/U+2011/U+2012/U+2015，導致同一題被當成兩題。
    # 這個字元集必須與 TS 端 question-identity.ts 的 PUNCT_RE 完全相同。
    s = re.sub(r"[，。？?、：:；;（）()「」『』【】\[\].,‐-―\-_/\\~｜|]", "", s)
    s = s.replace("臺", "台")
    return s


def clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    # openpyxl 讀數值欄會給 '1.0'
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return re.sub(r"[　]+", " ", s).strip()


def to_key(ans) -> str | None:
    a = clean(ans)
    if a in ("1", "2", "3", "4"):
        return KEYS[int(a) - 1]
    if a.upper() in KEYS:
        return a.upper()
    return None


def build_options(raw4: list[str]) -> list[dict]:
    """位置決定 key（選項1 恆為 A），空選項直接略過但不改變其他選項的字母。"""
    out = []
    for i, t in enumerate(raw4):
        t = clean(t)
        if t:
            out.append({"key": KEYS[i], "text": t})
    return out


# ---------------------------------------------------------------- load
def load_official() -> list[dict]:
    fh = open(OFFICIAL_CSV, encoding="utf-8")
    rows = list(csv.reader(fh))
    fh.close()
    items = []
    for r in rows[1:]:
        if len(r) < 7 or not clean(r[2]):
            continue
        no = int(float(r[0])) if clean(r[0]) else 0
        ans = to_key(r[1])
        opts = build_options(r[3:7])
        if ans is None or len(opts) < 2:
            continue
        items.append(
            {
                "id": f"tbfa-{no:03d}",
                "stem": clean(r[2]),
                "options": opts,
                "answer": ans,
                "explanation": clean(r[7]) if len(r) > 7 else "",
                "subject": SUBJ_LAW if no <= 240 else SUBJ_PRACTICE,
                "source_id": "tbfa-official-bank",
                "provenance": {
                    "source_type": "official_association_bank",
                    "original_no": no,
                },
                "tags": [],
            }
        )
    return items


def load_community() -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(COMMUNITY_XLSX, data_only=True)
    items = []
    seq = 0
    for ws in wb.worksheets:
        is_calc = ws.title == "計算題"
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[5]:
                continue
            ans = to_key(row[3])
            opts = build_options([row[6], row[7], row[8], row[9]])
            if ans is None or len(opts) < 2:
                continue
            no_raw = clean(row[1])
            try:
                no = int(float(no_raw)) if no_raw else 0
            except ValueError:
                no = 0
            seq += 1
            tags = []
            if is_calc:
                tags.append("計算題")
            if clean(row[0]):
                tags.append("易錯題")
            items.append(
                {
                    "id": f"comm-{seq:04d}",
                    "stem": clean(row[5]),
                    "options": opts,
                    "answer": ans,
                    "explanation": clean(row[10]),
                    "subject": SUBJ_LAW if 0 < no <= 240 else SUBJ_PRACTICE,
                    "source_id": "community-compilation-940",
                    "provenance": {
                        "source_type": "community_compilation",
                        "original_no": no or None,
                        "sheet": ws.title,
                    },
                    "tags": tags,
                }
            )
    return items


# ---------------------------------------------------------------- merge
def merge(official: list[dict], community: list[dict]):
    by_fp: dict[str, dict] = {}
    ordered: list[dict] = []
    for q in official:
        fp = norm_text(q["stem"])
        by_fp[fp] = q
        ordered.append(q)

    dup = 0
    conflicts = 0
    for q in community:
        fp = norm_text(q["stem"])
        base = by_fp.get(fp)
        if base is not None:
            dup += 1
            # 官方題保留為主；社群版若補得出解析或標記則併入
            if not base["explanation"] and q["explanation"]:
                base["explanation"] = q["explanation"]
                base["provenance"]["explanation_from"] = "community_compilation"
            for t in q["tags"]:
                if t not in base["tags"]:
                    base["tags"].append(t)
            base["provenance"]["also_in_community_compilation"] = True

            # 兩份來源對同一題給出不同答案 —— 這是「其中一方是錯的」的直接證據。
            #
            # 早期版本只保留官方版、把社群版默默丟掉。那等於銷毀了一條
            # 「這題有爭議」的訊號 —— 而這幾題正好是全題庫中最不該被照單全收的。
            # 現在把衝突記成一等欄位，並在介面上明確警示。
            # 注意：**這裡不裁決誰對。** 裁決需要查證，而本專案沒有做查證。
            if base["answer"] != q["answer"]:
                conflicts += 1
                # 用 kept/other 而非 official/community 命名：重複也可能發生在
                # 社群檔內部（同一題在兩個工作表答案不同），那種情況下兩邊都是
                # community，寫成 "official" 就是在說謊。
                base["provenance"]["answer_conflict"] = {
                    "kept": base["answer"],
                    "kept_source": base["provenance"]["source_type"],
                    "other": q["answer"],
                    "other_source": q["provenance"]["source_type"],
                }
                if "答案有爭議" not in base["tags"]:
                    base["tags"].append("答案有爭議")
            continue
        by_fp[fp] = q
        ordered.append(q)
    return ordered, dup, conflicts


SOURCES = [
    {
        "source_id": "tbfa-official-bank",
        "title": "票券商業務人員專業科目參考題庫（2 科各 240 題，合計 480 題）",
        "url": "https://www.tbfa.org.tw/BizTrain/biztrain_test.asp",
        "publisher": "中華民國票券金融商業同業公會",
        "authority": "official",
        "accessed_at": "2026-07-25",
        "note": "測驗委託單位公開釋出之參考題庫；含答案與法條依據。屬公會釋出之公開參考資料。",
    },
    {
        "source_id": "community-compilation-940",
        "title": "票券商業務人員考古題整理（Google 試算表，含「考題整理」與「計算題」兩工作表）",
        "url": "https://reurl.cc/Q6p0W5",
        "publisher": "不具名考生社群（經 PTT License 板考取心得文公開分享）",
        "authority": "community",
        "accessed_at": "2026-07-25",
        "note": "來源檔首列自述「有 1~2 題正解或選項描述有錯誤」。作者不具名，答案未經官方認可。",
    },
    {
        "source_id": "sfi-brochure-115",
        "title": "票券商業務人員專業科目測驗 電腦應試說明及操作手冊（115.05.01）",
        "url": "https://webline.sfi.org.tw/download/test_ftp/%E7%A5%A8%E5%88%B8%E7%B0%A1%E7%AB%A0.pdf",
        "publisher": "財團法人中華民國證券暨期貨市場發展基金會",
        "authority": "official",
        "accessed_at": "2026-07-25",
        "note": "考試制度資訊（題數／時間／合格標準／報名資格）之唯一權威來源。",
    },
    {
        "source_id": "moj-law-corpus",
        "title": "全國法規資料庫 — 票券金融管理法及其子法（17 部）",
        "url": "https://law.moj.gov.tw/LawClass/LawAll.aspx?pcode=G0380146",
        "publisher": "法務部",
        "authority": "official",
        "accessed_at": "2026-07-25",
        "note": "用於 check_law_citations.py 比對題目解析所引用之法條是否仍存在於現行條文。",
    },
]


def main() -> int:
    official = load_official()
    community = load_community()
    items, dup, conflicts = merge(official, community)

    by_subject: dict[str, int] = {}
    by_source: dict[str, int] = {}
    for q in items:
        by_subject[q["subject"]] = by_subject.get(q["subject"], 0) + 1
        st = q["provenance"]["source_type"]
        by_source[st] = by_source.get(st, 0) + 1

    dataset = {
        "meta": {
            "title": "票券商業務人員專業科目測驗 題庫（整合版）",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "schema_version": "1.0.0",
            "exam": {
                "name": "票券商業務人員專業科目測驗",
                "commissioner": "中華民國票券金融商業同業公會",
                "administrator": "財團法人中華民國證券暨期貨市場發展基金會",
                "legal_basis": "票券金融管理法第 12 條第 2 項授權訂定之「票券商負責人及業務人員管理規則」",
                "subjects": [
                    {"name": SUBJ_LAW, "questions": 50, "minutes": 60, "full_marks": 100},
                    {"name": SUBJ_PRACTICE, "questions": 50, "minutes": 60, "full_marks": 100},
                ],
                "passing_rule": "兩科總分合計達 140 分為合格，惟其中有任何 1 科分數低於 60 分者即屬不合格",
                "fee_twd": 1130,
                "brochure_version": "115.05.01",
                "brochure_source_id": "sfi-brochure-115",
            },
            "total_questions": len(items),
            "by_subject": by_subject,
            "by_source_type": by_source,
            "with_explanation": sum(1 for q in items if q["explanation"]),
            "deduped_against_official": dup,
            "answer_conflicts": {
                "count": conflicts,
                "note": (
                    "官方公會題庫與社群整理題庫對同一題給出不同答案的題數。"
                    "本專案**不裁決誰對** —— 這些題在介面上會標示「答案有爭議」並同時顯示兩方答案，"
                    "請自行查證現行法條。這是全題庫中最不該照單全收的一批題。"
                ),
            },
            "answer_verification": {
                "verified_against_official_key": 0,
                "note": (
                    "本題庫**沒有任何一題**的答案經過與現行法規逐條核對。官方公會題庫釋出年代較久，"
                    "部分答案可能已因法規修正而過時；社群整理題庫的來源檔更自述含有錯誤。"
                    "law_citation 欄位僅檢查「解析所引用的法條編號是否仍存在於現行條文」，"
                    "這是過時風險的**代理指標**，不代表答案正確。"
                ),
            },
        },
        "sources": SOURCES,
        "items": items,
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    fh = open(OUT, "w", encoding="utf-8")
    json.dump(dataset, fh, ensure_ascii=False, indent=1)
    fh.write("\n")
    fh.close()

    print(f"官方公會題庫          : {len(official)} 題")
    print(f"社群考古題整理        : {len(community)} 題（與官方重複 {dup} 題已併入）")
    print(f"  兩方答案衝突        : {conflicts} 題（已標記「答案有爭議」）")
    print(f"輸出總題數            : {len(items)} 題")
    for k, v in sorted(by_subject.items()):
        print(f"  {k}: {v}")
    for k, v in sorted(by_source.items()):
        print(f"  {k}: {v}")
    print(f"  含解析: {dataset['meta']['with_explanation']}")
    print(f"-> {OUT}  ({round(os.path.getsize(OUT)/1024,1)} KB)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
